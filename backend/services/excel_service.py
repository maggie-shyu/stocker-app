from __future__ import annotations

import threading
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from models.schemas import (
    AccountOverview,
    AccountOverviewHolding,
    CashflowCreate,
    CashflowRecord,
    StockLookup,
    TradeFinancials,
    TransactionCreate,
    TransactionRecord,
)
from services.calculator import calculate_trade_financials


WORKBOOK_WARNING = "contains invalid dependency definitions"


class ExcelService:
    _write_lock = threading.Lock()

    def __init__(self, workbook_path: str | Path):
        self.workbook_path = Path(workbook_path)
        self._stock_cache: list[StockLookup] | None = None

    def _load(self, *, data_only: bool = True):
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message=f".*{WORKBOOK_WARNING}.*")
            return load_workbook(self.workbook_path, data_only=data_only)

    @staticmethod
    def _date(value: Any) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            return date.fromisoformat(value[:10])
        raise ValueError(f"Unsupported date value: {value!r}")

    @staticmethod
    def _float(value: Any, default: float = 0) -> float:
        if value is None or value == "" or value == " ":
            return default
        return float(value)

    @staticmethod
    def _optional_float(value: Any) -> float | None:
        if value is None or value == "" or value == " ":
            return None
        return float(value)

    @staticmethod
    def _text(value: Any) -> str | None:
        if value is None:
            return None
        value = str(value).strip()
        return value or None

    def read_transactions(self) -> list[TransactionRecord]:
        ws = self._load(data_only=True)["交易紀錄"]
        rows: list[TransactionRecord] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if row[0] is None:
                continue
            rows.append(
                TransactionRecord(
                    id=row_idx,
                    date=self._date(row[0]),
                    action=row[1],
                    code=str(row[2]),
                    name=str(row[3]),
                    trade_type=row[4] or "一般",
                    buy_shares=self._optional_float(row[5]),
                    buy_price=self._optional_float(row[6]),
                    sell_shares=self._optional_float(row[7]),
                    sell_price=self._optional_float(row[8]),
                    current_price=self._float(row[9]),
                    raw_fee=self._float(row[10]),
                    discounted_fee=self._float(row[11]),
                    tax=self._float(row[12]),
                    amount=self._float(row[13]),
                    trade_cost=self._float(row[14]),
                    expense=self._float(row[15]),
                    income=self._float(row[16]),
                    reason=self._text(row[17]),
                    discount_rate=self._float(row[18]),
                )
            )
        return rows

    def append_transaction(
        self,
        payload: TransactionCreate,
        *,
        current_price: float = 0,
    ) -> TransactionRecord:
        discount_rate = self.get_commission_discount_rate()
        if payload.action == "買":
            shares, price = payload.buy_shares, payload.buy_price
        elif payload.action == "賣":
            shares, price = payload.sell_shares, payload.sell_price
        else:
            shares, price = None, None

        financials = calculate_trade_financials(
            action=payload.action,
            trade_type=payload.trade_type,
            code=payload.code,
            shares=shares,
            price=price,
            amount=payload.dividend_income if payload.action == "股利" else None,
            current_price=current_price,
            discount_rate=discount_rate,
        )
        with self._write_lock:
            wb = self._load(data_only=False)
            ws = wb["交易紀錄"]
            row_idx = self._first_empty_row(ws, start=3)
            values = [
                payload.date,
                payload.action,
                payload.code,
                payload.name,
                payload.trade_type,
                payload.buy_shares,
                payload.buy_price,
                payload.sell_shares,
                payload.sell_price,
                financials.current_price,
                financials.raw_fee,
                financials.discounted_fee,
                financials.tax,
                financials.amount,
                financials.trade_cost,
                financials.expense,
                financials.income,
                payload.reason,
                financials.discount_rate,
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row_idx, column=col, value=value)
            wb.save(self.workbook_path)
        return self._record_from_payload(row_idx, payload, financials)

    def delete_transaction(self, row_id: int) -> None:
        with self._write_lock:
            wb = self._load(data_only=False)
            ws = wb["交易紀錄"]
            if row_id < 3 or ws.cell(row=row_id, column=1).value is None:
                raise KeyError(row_id)
            ws.delete_rows(row_id, 1)
            wb.save(self.workbook_path)

    def read_cashflows(self) -> list[CashflowRecord]:
        ws = self._load(data_only=True)["出入金"]
        rows: list[CashflowRecord] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:
                continue
            rows.append(
                CashflowRecord(
                    id=row_idx,
                    date=self._date(row[0]),
                    deposit=self._float(row[1]),
                    withdrawal=self._float(row[2]),
                )
            )
        return rows

    def append_cashflow(self, payload: CashflowCreate) -> CashflowRecord:
        with self._write_lock:
            wb = self._load(data_only=False)
            ws = wb["出入金"]
            row_idx = self._first_empty_row(ws, start=2)
            for col, value in enumerate([payload.date, payload.deposit, payload.withdrawal], start=1):
                ws.cell(row=row_idx, column=col, value=value)
            wb.save(self.workbook_path)
        return CashflowRecord(
            id=row_idx,
            date=payload.date,
            deposit=payload.deposit,
            withdrawal=payload.withdrawal,
        )

    def get_commission_discount_rate(self) -> float:
        return self._float(self._load(data_only=True)["手續費"]["B1"].value)

    def set_commission_discount_rate(self, value: float) -> float:
        with self._write_lock:
            wb = self._load(data_only=False)
            wb["手續費"]["B1"] = value
            wb.save(self.workbook_path)
        return value

    def read_stocks(self) -> list[StockLookup]:
        if self._stock_cache is not None:
            return self._stock_cache
        ws = self._load(data_only=True)["股票代號表"]
        stocks: list[StockLookup] = []
        for code, name, *_ in ws.iter_rows(min_row=2, values_only=True):
            if code is None or name is None:
                continue
            stocks.append(StockLookup(code=str(code).strip(), name=str(name).strip()))
        self._stock_cache = stocks
        return stocks

    def search_stocks(self, query: str, *, limit: int = 20) -> list[StockLookup]:
        query = query.strip().lower()
        if not query:
            return self.read_stocks()[:limit]

        def score(stock: StockLookup) -> tuple[int, str]:
            code = stock.code.lower()
            name = stock.name.lower()
            if code == query or name == query:
                return (0, code)
            if code.startswith(query) or name.startswith(query):
                return (1, code)
            return (2, code)

        matches = [
            stock
            for stock in self.read_stocks()
            if query in stock.code.lower() or query in stock.name.lower()
        ]
        return sorted(matches, key=score)[:limit]

    def read_account_overview(self) -> AccountOverview:
        ws = self._load(data_only=True)["帳戶概況"]
        stocks_by_name = {stock.name: stock.code for stock in self.read_stocks()}
        holdings: list[AccountOverviewHolding] = []
        stock_market_value = 0.0
        cash_balance = 0.0
        account_value = self._float(ws["K2"].value)
        account_pnl = self._float(ws["L2"].value)
        account_pnl_rate = self._float(ws["M2"].value)

        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            name = self._text(row[0])
            if name is None:
                continue
            market_value = self._float(row[2])
            if name == "現金部位":
                cash_balance = market_value
                continue
            stock_market_value += market_value
            holdings.append(
                AccountOverviewHolding(
                    code=stocks_by_name.get(name),
                    name=name,
                    pnl_rate=self._optional_float(row[1]),
                    market_value=market_value,
                    weight=self._float(row[3]),
                )
            )
        return AccountOverview(
            account_value=account_value or stock_market_value + cash_balance,
            account_pnl=account_pnl,
            account_pnl_rate=account_pnl_rate,
            stock_market_value=stock_market_value,
            cash_balance=cash_balance,
            holdings=holdings,
        )

    @staticmethod
    def _first_empty_row(ws, *, start: int) -> int:
        row_idx = start
        while ws.cell(row=row_idx, column=1).value is not None:
            row_idx += 1
        return row_idx

    @staticmethod
    def _record_from_payload(
        row_idx: int,
        payload: TransactionCreate,
        financials: TradeFinancials,
    ) -> TransactionRecord:
        return TransactionRecord(
            id=row_idx,
            date=payload.date,
            action=payload.action,
            code=payload.code,
            name=payload.name,
            trade_type=payload.trade_type,
            buy_shares=payload.buy_shares,
            buy_price=payload.buy_price,
            sell_shares=payload.sell_shares,
            sell_price=payload.sell_price,
            current_price=financials.current_price,
            raw_fee=financials.raw_fee,
            discounted_fee=financials.discounted_fee,
            tax=financials.tax,
            amount=financials.amount,
            trade_cost=financials.trade_cost,
            expense=financials.expense,
            income=financials.income,
            reason=payload.reason,
            discount_rate=financials.discount_rate,
        )
