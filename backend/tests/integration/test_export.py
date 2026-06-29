import io
from datetime import date
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend.main import app
from backend.api import export as export_router
from backend.api.deps import get_current_user, get_ledger_store
from backend.models.api.auth import AuthenticatedUser
from backend.models.domain.ledger import CashflowRecord, TransactionRecord


client = TestClient(app)

FAKE_TX = TransactionRecord(
    id="tx-uuid-1",
    date=date(2025, 8, 6),
    action="買",
    code="0050",
    name="元大台灣50",
    trade_type="一般",
    buy_shares=1000,
    buy_price=200.0,
)
FAKE_CF = CashflowRecord(
    id="cf-uuid-1",
    date=date(2025, 1, 1),
    deposit=100000,
    withdrawal=0,
)


@pytest.fixture(autouse=True)
def override_deps():
    mock_svc = MagicMock()
    mock_svc.read_transactions.return_value = [FAKE_TX]
    mock_svc.read_cashflows.return_value = [FAKE_CF]
    mock_svc.replace_transactions.return_value = 1
    mock_svc.replace_cashflows.return_value = 1
    app.dependency_overrides[get_current_user] = lambda: AuthenticatedUser(
        id="test-user",
        email="tester@example.com",
    )
    app.dependency_overrides[get_ledger_store] = lambda: mock_svc
    yield
    app.dependency_overrides.clear()


def test_export_returns_xlsx():
    response = client.get("/api/export")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert "stocker_export.xlsx" in response.headers["content-disposition"]
    assert len(response.content) > 0


def test_export_has_two_sheets():
    workbook = openpyxl.load_workbook(io.BytesIO(client.get("/api/export").content))
    assert "交易紀錄" in workbook.sheetnames
    assert "出入金" in workbook.sheetnames


def test_export_transactions_sheet_has_header_and_data():
    workbook = openpyxl.load_workbook(io.BytesIO(client.get("/api/export").content))
    rows = list(workbook["交易紀錄"].iter_rows(values_only=True))
    assert rows[0][0] == "date"
    assert rows[1][2] == "0050"


def test_export_escapes_formula_like_cell_values():
    mock_svc = app.dependency_overrides[get_ledger_store]()
    mock_svc.read_transactions.return_value = [
        TransactionRecord(
            id="tx-uuid-2",
            date=date(2025, 8, 7),
            action="買",
            code="=cmd",
            name="@danger",
            trade_type="一般",
            buy_shares=1,
            buy_price=1.0,
            reason="+calc",
        )
    ]

    workbook = openpyxl.load_workbook(io.BytesIO(client.get("/api/export").content))
    rows = list(workbook["交易紀錄"].iter_rows(values_only=True))
    assert rows[1][2] == "'=cmd"
    assert rows[1][3] == "'@danger"
    assert rows[1][12] == "'+calc"


def test_import_reads_workbook_and_replaces_data():
    workbook = openpyxl.Workbook()
    transactions_sheet = workbook.active
    transactions_sheet.title = "交易紀錄"
    transactions_sheet.append([
        "date", "action", "code", "name", "trade_type",
        "buy_shares", "buy_price", "sell_shares", "sell_price",
        "dividend_shares", "dividend_price", "dividend_income", "reason"
    ])
    transactions_sheet.append([
        "2025-08-06", "買", "2330", "台積電", "一般",
        1000, 950, None, None, None, None, None, "imported"
    ])

    cashflow_sheet = workbook.create_sheet("出入金")
    cashflow_sheet.append(["date", "deposit", "withdrawal"])
    cashflow_sheet.append(["2025-01-01", 100000, 0])

    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/export/import",
        files={"file": ("stocker_import.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert response.json() == {
        "transactions_imported": 1,
        "cashflows_imported": 1,
    }

    mock_svc = app.dependency_overrides[get_ledger_store]()
    mock_svc.replace_transactions.assert_called_once()
    mock_svc.replace_cashflows.assert_called_once()


def test_import_rejects_oversized_upload(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(
        export_router,
        "get_settings",
        lambda: SimpleNamespace(
            import_max_upload_bytes=10,
            import_max_workbook_bytes=1000,
            import_max_rows_per_sheet=100,
        ),
    )

    workbook = openpyxl.Workbook()
    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/export/import",
        files={"file": ("stocker_import.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Workbook file is too large"


def test_import_rejects_excessive_row_count(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(
        export_router,
        "get_settings",
        lambda: SimpleNamespace(
            import_max_upload_bytes=5 * 1024 * 1024,
            import_max_workbook_bytes=25 * 1024 * 1024,
            import_max_rows_per_sheet=1,
        ),
    )

    workbook = openpyxl.Workbook()
    transactions_sheet = workbook.active
    transactions_sheet.title = "交易紀錄"
    transactions_sheet.append([
        "date", "action", "code", "name", "trade_type",
        "buy_shares", "buy_price", "sell_shares", "sell_price",
        "dividend_shares", "dividend_price", "dividend_income", "reason"
    ])
    transactions_sheet.append(["2025-08-06", "買", "2330", "台積電", "一般", 1000, 950, None, None, None, None, None, None])
    transactions_sheet.append(["2025-08-07", "買", "2317", "鴻海", "一般", 1000, 200, None, None, None, None, None, None])

    cashflow_sheet = workbook.create_sheet("出入金")
    cashflow_sheet.append(["date", "deposit", "withdrawal"])
    cashflow_sheet.append(["2025-01-01", 100000, 0])

    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)

    response = client.post(
        "/api/export/import",
        files={"file": ("stocker_import.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "交易紀錄 exceeds the maximum allowed row count"
