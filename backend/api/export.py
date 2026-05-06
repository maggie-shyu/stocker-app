import io
import zipfile
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from backend.api.deps import get_ledger_store
from backend.config import get_settings
from backend.domain.ledger.interfaces import LedgerStore
from backend.models.api.ledger import CashflowCreate, TransactionCreate


router = APIRouter(prefix="/api/export", tags=["export"])


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return date.fromisoformat(value[:10])
    raise ValueError(f"Unsupported date value: {value!r}")


def _optional_float(value):
    if value is None or value == "":
        return None
    return float(value)


def _safe_excel_cell(value):
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    if stripped and stripped[0] in ("=", "+", "-", "@"):
        return f"'{value}"
    if value and value[0] in ("\t", "\r", "\n"):
        return f"'{value}"
    return value


@router.get("")
def export_data(ledger_store: LedgerStore = Depends(get_ledger_store)):
    workbook = Workbook()

    transactions_sheet = workbook.active
    transactions_sheet.title = "交易紀錄"
    transactions_sheet.append(
        [
            "date",
            "action",
            "code",
            "name",
            "trade_type",
            "buy_shares",
            "buy_price",
            "sell_shares",
            "sell_price",
            "dividend_income",
            "reason",
        ]
    )
    for tx in ledger_store.read_transactions():
        transactions_sheet.append(
            [
                str(tx.date),
                _safe_excel_cell(tx.action),
                _safe_excel_cell(tx.code),
                _safe_excel_cell(tx.name),
                _safe_excel_cell(tx.trade_type),
                tx.buy_shares,
                tx.buy_price,
                tx.sell_shares,
                tx.sell_price,
                tx.income if tx.action == "股利" else None,
                _safe_excel_cell(tx.reason),
            ]
        )

    cashflow_sheet = workbook.create_sheet("出入金")
    cashflow_sheet.append(["date", "deposit", "withdrawal"])
    for cashflow in ledger_store.read_cashflows():
        cashflow_sheet.append(
            [
                str(cashflow.date),
                cashflow.deposit,
                cashflow.withdrawal,
            ]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stocker_export.xlsx"},
    )


@router.post("/import")
async def import_data(
    file: UploadFile = File(...),
    ledger_store: LedgerStore = Depends(get_ledger_store),
):
    settings = get_settings()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    try:
        content = await file.read()
        if len(content) > settings.import_max_upload_bytes:
            raise HTTPException(status_code=400, detail="Workbook file is too large")
        if not zipfile.is_zipfile(io.BytesIO(content)):
            raise HTTPException(status_code=400, detail="Uploaded file is not a valid .xlsx workbook")

        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            total_uncompressed_bytes = sum(item.file_size for item in archive.infolist())
        if total_uncompressed_bytes > settings.import_max_workbook_bytes:
            raise HTTPException(status_code=400, detail="Workbook content is too large")

        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        if "交易紀錄" not in workbook.sheetnames or "出入金" not in workbook.sheetnames:
            raise HTTPException(status_code=400, detail="Workbook must contain 交易紀錄 and 出入金 sheets")
        if workbook["交易紀錄"].max_row - 1 > settings.import_max_rows_per_sheet:
            raise HTTPException(status_code=400, detail="交易紀錄 exceeds the maximum allowed row count")
        if workbook["出入金"].max_row - 1 > settings.import_max_rows_per_sheet:
            raise HTTPException(status_code=400, detail="出入金 exceeds the maximum allowed row count")

        transaction_rows = list(workbook["交易紀錄"].iter_rows(min_row=2, values_only=True))
        cashflow_rows = list(workbook["出入金"].iter_rows(min_row=2, values_only=True))

        transactions = [
            TransactionCreate(
                date=_parse_date(row[0]),
                action=row[1],
                code=str(row[2]),
                name=str(row[3]),
                trade_type=row[4] or "一般",
                buy_shares=_optional_float(row[5]),
                buy_price=_optional_float(row[6]),
                sell_shares=_optional_float(row[7]),
                sell_price=_optional_float(row[8]),
                dividend_income=_optional_float(row[9]),
                reason=str(row[10]).strip() if row[10] is not None else None,
            )
            for row in transaction_rows
            if row[0] is not None
        ]
        cashflows = [
            CashflowCreate(
                date=_parse_date(row[0]),
                deposit=float(row[1] or 0),
                withdrawal=float(row[2] or 0),
            )
            for row in cashflow_rows
            if row[0] is not None
        ]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Failed to parse workbook") from exc

    transaction_count = ledger_store.replace_transactions(transactions)
    cashflow_count = ledger_store.replace_cashflows(cashflows)
    return {
        "transactions_imported": transaction_count,
        "cashflows_imported": cashflow_count,
    }
